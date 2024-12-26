import os
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl import Workbook

class ExcelOperations:

    def create_or_append_excel(self, file_name, rows_to_add):
        """
        Creates a new Excel workbook if 'file_name' does not exist.
        If 'file_name' exists, appends data row-by-row to the existing workbook.
        
        :param file_name:   The Excel file name/path (e.g., "data.xlsx")
        :param rows_to_add: A list of lists, where each inner list is a row to append
        """

        # Check if the Excel file already exists
        if os.path.exists(file_name):
            print(f"'{file_name}' already exists. Loading and appending rows...")
            # Load the existing workbook
            wb = load_workbook(file_name)
            ws = wb.active  # Use the active sheet (or specify by name)
        else:
            print(f"'{file_name}' does not exist. Creating a new workbook...")
            # Create a new workbook
            wb = Workbook()
            ws = wb.active

        # Append rows (each item in rows_to_add is a list representing one row)
        ws.append(rows_to_add)
        
        # Save the workbook (overwrites if file already exists)
        wb.save(file_name)
        print(f"Data successfully appended to '{file_name}'")